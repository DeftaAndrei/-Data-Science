import pandas as pd 
import numpy as np
from collections import Counter
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_text(text):
    if isinstance(text, str):
        text = re.sub(r'[^\w\s]', ' ', text.lower())
        return re.sub(r'\s+', ' ', text).strip()
    return ''

# Citim datele în liniște
taxonomy_df = pd.read_csv('1.csv')
companies_df = pd.read_csv('ml_insurance_challenge.csv')

# Redenumim coloanele pentru a se potrivi cu formatul dorit
column_mapping = {
    'description': 'description',
    'business_tags': 'business_tags',
    'sector': 'sector',
    'category': 'category',
    'niche': 'niche'
}

# Ne asigurăm că avem coloanele corecte
for old_col, new_col in column_mapping.items():
    if old_col in companies_df.columns:
        companies_df = companies_df.rename(columns={old_col: new_col})

# Creăm un nou workbook
wb = Workbook()
ws = wb.active
ws.title = 'Sheet1'

# Adăugăm headerele
headers = ['description', 'business_tags', 'sector', 'category', 'niche']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    
    # Formatare header
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='0F243E', end_color='0F243E', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

# Adăugăm datele
for row_idx, row in enumerate(companies_df.values, 2):
    for col_idx, value in enumerate(row[:5], 1):  # Luăm doar primele 5 coloane
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = str(value) if pd.notna(value) else ''
        
        # Formatare celule de date
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Setăm lățimea coloanelor
column_widths = {
    'A': 30,  # description
    'B': 25,  # business_tags
    'C': 15,  # sector
    'D': 20,  # category
    'E': 20   # niche
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Setăm înălțimea rândurilor
ws.row_dimensions[1].height = 20  # Header height
for row in range(2, len(companies_df) + 2):
    ws.row_dimensions[row].height = 30  # Data rows height

# Adăugăm borduri pentru toate celulele
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=len(companies_df)+1, min_col=1, max_col=5):
    for cell in row:
        cell.border = thin_border

# Salvăm workbook-ul
wb.save('Rezolvare Company Classifier.xlsx')

print("Excel creat cu succes!")





